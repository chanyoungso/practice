# in openpyxl indices of row and column start with one not zero
import openpyxl
import numpy
import math
import scipy.stats as stats
import itertools



class UnitableAnalyzer( ):
    
    # declare workbook as a global variable, whiel worksheet is declared as a local variable
    unitable_workbook = openpyxl.Workbook()
    unitable_file_name = ''    
    src_sheet_name = ''
    
    bivariate_workbook = openpyxl.Workbook()
    bivariate_file_name = ''
    dst_sheet_name = ''
    
    # set starting and ending points
    # data is stored along the column for each question (or item) which is aligned along the row
    # the first row is used for the indicator of question number(variable name)
    num_of_entry = 1
    num_of_vars = 1
    num_of_rows = 1
    start_row = 2 # common for source and destination
    start_col = 2 # common for source and destination
    end_row = start_row + num_of_entry - 1
    end_col = start_col + num_of_vars - 1 # only for source
    
    # initialize data list    
    src_var_name_list = []
    dst_var_name_list = []
    entry_list = []
    data_list = [[], []] # for bivariate analysis only    
    var_name_combination_list = []
    
        
    def __init__( self, unitable_file_name, bivariate_file_name ): 
        
        self.open_file( unitable_file_name, bivariate_file_name)
        self.unitable_file_name = unitable_file_name
        self.bivariate_file_name = bivariate_file_name
        
        return None
    
    
    # externally called
    def set_target_sheet( self, sheet_name ):
        
        self.src_sheet_name = sheet_name
        worksheet = self.unitable_workbook.get_sheet_by_name( sheet_name )
             
        self.num_of_vars = worksheet.max_column - 1
        self.num_of_rows = worksheet.max_row - 1
        
        self.end_row = self.start_row + self.num_of_rows - 1
        self.end_col = self.start_col + self.num_of_vars - 1

        return True
    
    
    def make_combination ( self, original_list ): 
        
        combination_list = []
        for cur_comb in itertools.combinations( original_list, 2 ):
            combination_list.append( cur_comb )           
            
        return combination_list
    
    
    def add_dst_sheet( self, sheet_name ):
        
        self.bivariate_workbook.create_sheet( sheet_name )
        self.dst_sheet_name = sheet_name
        worksheet = self.get_dst_sheet()
        worksheet.title = sheet_name
        worksheet.sheet_view.showGridLines = False
        
        return True

    
    def get_src_sheet( self ):
        
        worksheet = self.unitable_workbook.get_sheet_by_name( self.src_sheet_name )
        
        return worksheet
    
    
    def get_dst_sheet( self ):
        
        worksheet = self.bivariate_workbook.get_sheet_by_name( self.dst_sheet_name )
        
        return worksheet
    
    
    # functions to control excel file
    def open_file( self, unitable_file_name, bivariate_file_name ):
        
        self.unitable_workbook = openpyxl.load_workbook( unitable_file_name )        
        new_workbook = openpyxl.Workbook()
        new_workbook.save( bivariate_file_name )
        self.bivariate_workbook = openpyxl.load_workbook( bivariate_file_name )
        
        return True    
    
    
    # externally called
    def save_to_file( self ):        
        
        self.bivariate_workbook.save( self.bivariate_file_name )
        
        return True 
    
    
    # externally called
    def close_file( self ):
        
        self.unitable_workbook.close()
        self.bivariate_workbook.close()
        
        return True        
    
    
    def initialize_data( self ):
        
        self.src_var_name_list = []
        self.dst_var_name_list = []
        self.entry_list = []
        self.data_list = [[], []]        
        self.var_name_combination_list = []        

        return True
    
    def make_sole_list( self ):
        
        self.initialize_data()
        worksheet = self.get_src_sheet()
        
        # read entry from the first column
        for col in worksheet.iter_cols( max_col = 1, min_row = self.start_row, max_row = self.end_row ):
            for cell in col:                
                self.entry_list.append( cell.value )
           
        self.entry_list = list( set ( self.entry_list ) )
        self.entry_list.sort()
        self.num_of_entry = len ( self.entry_list )
                
        # read variable names from the first row
        for row in worksheet.iter_cols( max_row = 1, min_col = self.start_col, max_col = self.end_col ):
            for cell in row:                
                self.src_var_name_list.append( cell.value )
                
        self.num_of_vars = len( self.src_var_name_list )        
        
        return True
    
    
    def load_comb_data( self, src_var_names ):
        
        worksheet = self.get_src_sheet()
        
        self.dst_var_name_list = []
        self.data_list = [[],[]]
        self.dst_var_name_list = list ( src_var_names )
        
        max_repeat_time = 0
        
        # read data from two columns        
        for cur_index in range( 0, 2 ):
            
            src_col = 1
            for cur_col in range( 2, self.end_col + 1 ):  
                
                if worksheet.cell( row = 1, column = cur_col ).value == src_var_names[cur_index]:                                        
                    src_col = cur_col
            
            for col in worksheet.iter_cols( min_row = self.start_row, max_row = self.end_row, min_col = src_col, max_col = src_col ):                    
                for cell in col:
                    self.data_list[cur_index].append( cell.value )
                    
            cur_repeat_time = self.get_dup_time( self.data_list[cur_index], self.num_of_entry )
            self.data_list[cur_index] = self.data_list[cur_index][0:(self.num_of_entry * cur_repeat_time)]

            if max_repeat_time < cur_repeat_time:
                max_repeat_time = cur_repeat_time                     
                
        return max_repeat_time
    
    
    # externally called
    def write_to_bivariate_file( self ):
                
        self.make_sole_list()        
        self.var_name_combination_list = self.make_combination( self.src_var_name_list )

        for cur_var_names in self.var_name_combination_list:            
            repeat_time = self.load_comb_data( cur_var_names )
            sheet_name = list( cur_var_names )[0] + 'x' + list( cur_var_names )[1]            
            self.add_dst_sheet( sheet_name )
            self.write_to_dst_sheet( repeat_time )
            
        self.bivariate_workbook.remove_sheet( self.bivariate_workbook.get_sheet_by_name( 'Sheet' ) )
                                             
        return True
    
    
    def get_dup_time( self, data_list, length_of_data ):
        
        repeat_time = 1
                
        data_end = 0
        for cur_index in range(0, len( data_list ) ):            
            cur_pos = cur_index + 1
            cur_data = data_list[cur_index]
            if ( not cur_data is None ) and ( data_end < cur_pos ):
                data_end = cur_pos
        
        repeat_time = math.ceil( data_end / length_of_data )
        
        if repeat_time == 0:
            repeat_time = 1
        
        return repeat_time
    
    
    def write_to_dst_sheet( self, repeat_time ):        
        
        worksheet = self.get_dst_sheet()        
        worksheet.cell( row = 1, column = 1 ).value = 'No.'
     
        # write var names
        for cur_index in range( 0, 2 ):
            worksheet.cell( row = 1, column = self.start_col + cur_index  ).value = self.dst_var_name_list[cur_index]
        
        # write entry names
        for cur_time in range( 0, repeat_time ):
            for cur_index in range( 0, self.num_of_entry ):
                worksheet.cell( row = self.start_row + cur_index + ( self.num_of_entry * cur_time ), column = 1 ).value = self.entry_list[cur_index]
                
        # write values
        for cur_col in range( 0, 2 ):
                       
            cur_dup_size = self.get_dup_time( self.data_list[cur_col], self.num_of_entry )            
            sequence = []
            
            quotient = repeat_time // cur_dup_size
            remainder = repeat_time % cur_dup_size
            
            sequence += [ cur_dup_size ] * quotient
            sequence.append( remainder )
            
            cur_row = self.start_row            
            for cur_step in sequence:                
                for cur_index in range( 0, ( self.num_of_entry * cur_step ) ):                        
                    worksheet.cell( row = cur_row, column = 2 + cur_col ).value = self.data_list[cur_col][cur_index]
                    cur_row += 1
                
        return True
    
    
    # checking functions
    def is_data_loaded( self ):        
        
        # to be modified ( ... )
        
        return True
  
    
    def is_target_sheet_set( self ):
        
        # to be modified ( ... )
        
        return True
            


class BivariateAnalyzer:
    
    # declare workbook as a global variable, whiel worksheet is declared as a local variable
    bivariate_workbook = openpyxl.Workbook()
    bivariate_file_name = ''    
    bivariate_sheet_name = ''
    summary_sheet_name = 'Summary'        
    
    # set starting and ending points
    # data is stored along the column for each question (or item) which is aligned along the row
    # the first row is used for the indicator of question number
    num_of_entry = 1
    num_of_vars = 1
    start_row = 2
    start_col = 2
    end_row = start_row + num_of_entry - 1
    end_col = start_col + num_of_vars - 1 

    sum_start_row = 3
    sum_start_col = 5       
    
    # initialize data list    
    var_name_list = []
    entry_list = []
    data_list = [[], []] # for bivariate analysis only
    data_set = [ var_name_list, entry_list, data_list ] # to be changed    
    data_type_list = []

    # data type constants
    type_nom = 'nominal'
    type_ord = 'ordinal'
    type_cat = 'categorical'
    type_int = 'interval'
    type_rat = 'ratio'
    type_num = 'numeric'

    # result type constants
    type_p_value = 'p-value'
    type_pearsons_r = 'pearson\'s r'
    type_anova = 'anova'
    type_boxplot = 'boxplot'
    type_f_ratio = 'f_ratio'
    
    # color codes
    color_orange_red = '8FAADC'
    color_heavy_blue = 'F4B183'
    color_light_blue = '5B9BD5'
    color_light_green = 'E2EFDA'
    color_light_yellow = 'FFF2CC'
    color_light_bright_blue = 'D9E1F2'
    color_white = 'FFFFFF'
    
    
    def __init__( self, bivariate_file_name ):
        
        self.open_file( bivariate_file_name )
        self.bivariate_file_name = bivariate_file_name          
        
        return None
        
    
    def set_target_sheet( self, bivariate_sheet_name ):
        
        self.bivariate_sheet_name = bivariate_sheet_name
        worksheet = self.bivariate_workbook.get_sheet_by_name( bivariate_sheet_name )
             
        # check the length of columns, which is equal to the number of interviewees plus one (for the indicator of question number)                
        self.num_of_vars = 2 # to avoid unintended error
        self.num_of_entry = worksheet.max_row - 1    # be cautious     
        
        self.end_row = self.start_row + self.num_of_entry - 1
        self.end_col = self.start_col + self.num_of_vars - 1

        self.load_from_file()        
        
        return True
    
    
    def get_current_sheet( self ):        
        
        if not self.is_target_sheet_set() and self.is_data_loaded():
            return None
        
        worksheet = self.bivariate_workbook.get_sheet_by_name( self.bivariate_sheet_name )
        return worksheet

    
    def get_sheet_name_list( self ):
        
        return self.bivariate_workbook.get_sheet_names()
    

    def get_cell_address( self, col, row, absolute = False ):
        
        return openpyxl.utils.get_column_letter( col ) + str( row )
    
    
    def get_range_address( self,  start_col, start_row, end_col, end_row, absolute = False  ):
        
        return self.get_cell_address( start_col, start_row ) + ':' + self.get_cell_address( end_col, end_row )
    
    
    def initialize_data( self ):
        
        self.var_name_list = []
        self.entry_list = []
        self.data_list = [[], []]
        self.data_type_list = []        
        
        return True        
  
                  
    def load_from_file( self ):
        
        if not self.is_target_sheet_set() and self.is_data_loaded():
            return True
            
        self.initialize_data()
        worksheet = self.get_current_sheet()
    
        # read entry from the first column
        for col in worksheet.iter_cols( max_col = 1, min_row = self.start_row, max_row = self.end_row ):
            for cell in col:
                self.entry_list.append( cell.value )
                
        # read variable names from the first row
        for row in worksheet.iter_cols( max_row = 1, min_col = self.start_col, max_col = self.end_col ):
            for cell in row:
                self.var_name_list.append( cell.value )

        # read data from two columns
        cur_data_index = 0
        for col in worksheet.iter_cols( min_col = self.start_col, max_col = self.end_col,\
                                       min_row = self.start_row, max_row = self.end_row ):
            for cell in col:                    
                self.data_list[cur_data_index].append( cell.value )                    
            cur_data_index += 1
    
        self.rectify_none()
        self.scan_data_type()
        
        return True
 
    
    def rectify_none( self ):
        
        none_index_list = []        
        
        # only two columns
        for cur_index in range(0, 2):                        
            none_index_list += [index for index, value in enumerate( self.data_list[cur_index] ) if value == None]
            
        none_index_list = list( set( none_index_list ) )
        none_index_list.sort( reverse = True )
        
        for cur_index in none_index_list:
            del( self.entry_list[cur_index] )
            for cur_data_index in range(0, 2):
                del( self.data_list[cur_data_index][cur_index] )        
        
        return True
    
    
    def scan_data_type( self ):
        
        self.data_type_list = []        
        
        for cur_index in range( 0, 2 ):            
            self.data_type_list.append( self.check_data_type( self.data_list[cur_index] ) )
            
        return True
    
    
    # functions to control excel file
    def open_file( self, bivariate_file_name):
        
        self.bivariate_workbook = openpyxl.load_workbook( bivariate_file_name )
        
        return True
        
    
    def save_to_file( self ):

        self.bivariate_workbook.active = self.bivariate_workbook.sheetnames.index( self.summary_sheet_name )
        self.bivariate_workbook.save( self.bivariate_file_name )        
        
        return True
 
    
    def close_file( self ):
        
        self.bivariate_workbook.close()         
        
        return True
   
       
    # checking functions
    def is_data_loaded( self ):        
        
        # to be modified ( ... )
        
        return True
  
    
    def is_target_sheet_set( self ):
        
        # to be modified ( ... )
        
        return True
    
    
    def check_data_type( self, target_list ):
        
        data_type = ''
        
        for data in target_list:
                           
            if str( data ).isdigit() or self.is_float( str( data ) ) :
                data_type = self.type_num                
            else:
                data_type = self.type_nom
                break
                
        return data_type
    
    
    def is_float( self, data ):
        
        try:
            float( data )
            return True
        
        except ValueError:            
            return False
       
        
    def wrapper( self, func, args):        
        return func( *args )
    
    
    # ===== functions to get statistic values =====
    
    def get_pearsons_r( self, a_list, b_list ):        
        
        if not self.is_target_sheet_set() and self.is_data_loaded():
            return 0
        
        pearsons_r_matrix = numpy.corrcoef( a_list, b_list )          
        pearsons_r = pearsons_r_matrix[0][1] # to be changed            
        
        return pearsons_r
  
    
    def get_cramers_v( self, a_list, b_list ):
        
        sole_a_list = list( set( a_list ) )
        sole_a_list.sort( reverse = False )
        
        sole_b_list = list( set( b_list ) )
        sole_b_list.sort( reverse = False )
        
        num_of_a = len( sole_a_list )
        num_of_b = len( sole_b_list )
        
        total_number = min( len( a_list ), len( b_list ) )        
        pearson_chi_square = self.get_pearson_chi_square( a_list, b_list )
        phi_coefficient = pearson_chi_square / total_number
        cramers_v = math.sqrt( phi_coefficient / min( num_of_a, num_of_b )  ) 
        
        return cramers_v        
  
    
    def get_spearmans_rho( self, data_list ):
        
        v = 0
        return v
    
    
    def get_hypothesis_test_result( self, test_value, critical_value, reject_region = True ):
        
        result = ''
        
        if test_value >= critical_value:
            result = 'Alternative hypothesis accepted.'
            
        else:
            result = 'Alternative hypothesis rejected.'
            
        return result
    
    
    def get_pearson_chi_square( self, a_list, b_list ):
        
        if not self.is_target_sheet_set() and self.is_data_loaded():
            return 0
        
        sole_a_list = list( set( a_list ) )
        sole_a_list.sort( reverse = False )
        
        sole_b_list = list( set( b_list ) )
        sole_b_list.sort( reverse = False )
        
        num_of_a = len( sole_a_list )
        num_of_b = len( sole_b_list )
        
        observed_freq_list = []
        expected_freq_list = []
        
        a_prob_sum_list = []
        b_prob_sum_list = []

        total_observation = 1
        if len( a_list ) == len( b_list ):            
            total_observation =  len( a_list )
        
        couple_list = list( zip( a_list, b_list ) )
            
        # observed frequency
        for cur_b_index in range( 0, num_of_b ):
            
            a_observed_freq_list = []
            for cur_a_index in range( 0, num_of_a ):
                 a_observed_freq_list.append( couple_list.count( ( sole_a_list[cur_a_index], sole_b_list[cur_b_index] ) ) ) # to be changed
                
            observed_freq_list.append( a_observed_freq_list )
                
        # observed probability sum
        for cur_a_index in range( 0, num_of_a ):
            a_prob_sum_list.append( sum( observed_freq_list[cur_b_index][cur_a_index] for cur_b_index in range( 0, num_of_b ) ) / total_observation )            
            
        for cur_b_index in range( 0, num_of_b ):
            b_prob_sum_list.append( sum( observed_freq_list[cur_b_index] ) / total_observation )
                    
        # expected_freq        
        for cur_b_index in range( 0, num_of_b ):
            
            a_expected_freq_list = []
            for cur_a_index in range( 0, num_of_a ):                
                a_expected_freq_list.append( total_observation * a_prob_sum_list[cur_a_index] * b_prob_sum_list[cur_b_index] )
            
            expected_freq_list.append( a_expected_freq_list )
       
        # consolidate each list              
        temp_list = []
        for cur_list in observed_freq_list:
            temp_list += cur_list
        
        observed_freq_list = temp_list
                
        temp_list = []
        for cur_list in expected_freq_list:
            temp_list += cur_list
        
        expected_freq_list = temp_list
        
        pearson_chi_square = stats.chisquare( f_obs = observed_freq_list, f_exp = expected_freq_list )[0] # to be changed
                
        return pearson_chi_square
    
    
    # ===== functions to get excel formulae =====
    
    def get_freq_formula( self, a_ref_range, a_criteria_cell, b_ref_range, b_criteria_cell ):        
        formula = '=_xlfn.COUNTIFS(' + a_ref_range + ',' + a_criteria_cell + ',' + b_ref_range + ',' + b_criteria_cell + ')'                
        return formula
        
    def get_prob_formula( self, dividend_cell, divisor_cell ):
        formula = '=' + dividend_cell + '/' + divisor_cell
        return formula    
    
    def get_exp_freq_formula( self, total_cell, a_prob_cell, b_prob_cell ):
        formula = '=' + total_cell + '*' + a_prob_cell + '*' + b_prob_cell
        return formula

    def get_chi_square_formula( self, freq_cell, exp_freq_cell ):
        formula = '=_xlfn.IFERROR(((' + freq_cell + '-' + exp_freq_cell + ')^2)/' + exp_freq_cell + ',0)'
        return formula
    
    def get_sum_formula( self, ref_range ):
        formula = '=_xlfn.SUM(' + ref_range + ')'
        return formula
    
    
    # ===== functions for summary =====
    
    def add_summary_sheet ( self ):
        
        if self.summary_sheet_name in self.bivariate_workbook.sheetnames:
            self.bivariate_workbook.remove_sheet( self.summary_sheet_name )
                
        self.bivariate_workbook.create_sheet( self.summary_sheet_name )        
        worksheet = self.bivariate_workbook.get_sheet_by_name( self.summary_sheet_name )
        worksheet.sheet_view.showGridLines = False
        
        worksheet.cell( row = 1, column = 1 ).value = 'No.'
        worksheet.cell( row = 1, column = 2 ).value = 'Question'
        worksheet.cell( row = 2, column = 2 ).value = 'Type'
        
        bivariate_sheet_list = self.bivariate_workbook.sheetnames
        num_of_sheets = len( bivariate_sheet_list )
        var_names_list = []
                
        for cur_index in range( 0, num_of_sheets ):            
            var_names_comb = bivariate_sheet_list[cur_index].split('x')            
            for cur_var in var_names_comb:
                if not cur_var in var_names_list and not cur_var == self.summary_sheet_name:
                    var_names_list.append( cur_var )
        
        num_of_vars = len( var_names_list )
        
        for cur_index in range( 0, num_of_vars ):            
            worksheet.cell( row = 1, column = self.sum_start_col + cur_index ).value = var_names_list[cur_index]
            worksheet.cell( row = self.sum_start_row + cur_index, column = 1 ).value = var_names_list[cur_index]
            worksheet.cell( row = self.sum_start_row + cur_index, column = self.sum_start_col + cur_index ).value =  1
            worksheet.cell( row = self.sum_start_row + cur_index, column = self.sum_start_col + cur_index ).number_format = '0.00'        
        
        worksheet.merge_cells( start_row = 1, start_column = 1, end_row = 2, end_column = 1 )
        worksheet.merge_cells( start_row = 1, start_column = 2, end_row = 1, end_column = self.sum_start_col - 1 )
        worksheet.merge_cells( start_row = 2, start_column = 2, end_row = 2, end_column = self.sum_start_col - 1 )
        
        worksheet.cell( row = 1, column = 1 ).alignment = openpyxl.styles.Alignment( horizontal = 'center', vertical = 'center' )
        worksheet.cell( row = 1, column = 2 ).alignment = openpyxl.styles.Alignment( horizontal = 'center', vertical = 'center' )
        worksheet.cell( row = 2, column = 2 ).alignment = openpyxl.styles.Alignment( horizontal = 'center', vertical = 'center' )
        
        data_end_row = worksheet.max_row
        
        worksheet.cell( row = data_end_row + 2, column = self.sum_start_col ).fill = openpyxl.styles.fills.PatternFill( patternType='solid',
                      fgColor = openpyxl.styles.colors.Color( rgb = self.color_light_bright_blue ) )
        worksheet.cell( row = data_end_row + 3, column = self.sum_start_col ).fill = openpyxl.styles.fills.PatternFill( patternType='solid',
                      fgColor = openpyxl.styles.colors.Color( rgb = self.color_light_green ) )
        worksheet.cell( row = data_end_row + 4, column = self.sum_start_col ).fill = openpyxl.styles.fills.PatternFill( patternType='solid',
                      fgColor = openpyxl.styles.colors.Color( rgb = self.color_light_yellow ) )
        
        worksheet.cell( row = data_end_row + 2, column = self.sum_start_col + 1 ).value = 'Cross table'
        worksheet.cell( row = data_end_row + 3, column = self.sum_start_col + 1 ).value = 'Boxplot + ANOVA'
        worksheet.cell( row = data_end_row + 4, column = self.sum_start_col + 1 ).value = 'Correlation'

        return True                        
                
    
    def write_summary( self, var_name_list, data_type_list, result_value, result_type ):        
         
        # value a will be put in column index line, value b will be put in row index line
        worksheet = self.bivariate_workbook.get_sheet_by_name( self.summary_sheet_name )
         
        # finding coordinates
        target_row = self.sum_start_row
        target_col = self.sum_start_col
        x_row = self.sum_start_row
        x_col = self.sum_start_col
        data_end_row = worksheet.max_row
        data_end_col = worksheet.max_column        
        
        for cur_col in range( self.sum_start_col, data_end_col + 1 ):
            if worksheet.cell( row = 1, column = cur_col ).value == var_name_list[0]:
                target_col = cur_col
            if worksheet.cell( row = 1, column = cur_col ).value == var_name_list[1]:
                x_col = cur_col
         
        for cur_row in range( self.sum_start_row, data_end_row + 1 ):
            if worksheet.cell( row = cur_row, column = 1 ).value == var_name_list[1]:
                target_row = cur_row
            if worksheet.cell( row = cur_row, column = 1 ).value == var_name_list[0]:
                x_row = cur_row
        
        # result value filling
        worksheet.cell( row = target_row, column = target_col ).value = result_value
        worksheet.cell( row = target_row, column = target_col ).number_format = '0.00'
        worksheet.cell( row = target_row, column = target_col ).hyperlink = '#\'' + var_name_list[0] + 'x' + var_name_list[1] + '\'!' + 'A1'        
        
        filling_color_code = self.color_white
        if result_type == self.type_p_value:
            filling_color_code = self.color_light_bright_blue
            
        elif result_type == self.type_anova:
            filling_color_code = self.color_light_green
            
        elif result_type == self.type_pearsons_r:
            filling_color_code = self.color_light_yellow
                
        filling_style = openpyxl.styles.fills.PatternFill( patternType='solid', fgColor = openpyxl.styles.colors.Color( rgb = filling_color_code ) )
        worksheet.cell( row = target_row, column = target_col ).fill = filling_style
                
        # symmetry x filling                 
        worksheet.cell( row = x_row, column = x_col ).value = 'x'
        worksheet.cell( row = x_row, column = x_col ).alignment = openpyxl.styles.Alignment( horizontal = 'center', vertical = 'center' )
        
        # data type filling
        worksheet.cell( row = 2, column = target_col ).value = data_type_list[0]
        worksheet.cell( row = x_row, column = self.sum_start_col - 1 ).value = data_type_list[0]        
        worksheet.cell( row = target_row, column = self.sum_start_col - 1 ).value = data_type_list[1]
        worksheet.cell( row = 2, column = x_col ).value = data_type_list[1]
        
        return True
    
    
    # ===== analysis functions =====
    
    def analyze_current_sheet( self ):
        
        result_value = 0
        result_type = ''        
        if not self.summary_sheet_name in self.bivariate_workbook.sheetnames:            
            self.add_summary_sheet()
       
        if self.data_type_list[0] == self.type_num and self.data_type_list[1] == self.type_num:
            result_value = self.analyze_correl()
            result_type = self.type_pearsons_r
        
        elif ( self.data_type_list[0] == self.type_nom and self.data_type_list[1] == self.type_num ) or\
             ( self.data_type_list[0] == self.type_num and self.data_type_list[1] == self.type_nom ):
            result_value = self.analyze_anova()
            result_type = self.type_anova
        
        elif self.data_type_list[0] == self.type_nom and self.data_type_list[1] == self.type_nom:
            result_value = self.analyze_cross_table()
            result_type = self.type_p_value
        
        else:
            pass
        
        self.write_summary( self.var_name_list, self.data_type_list, result_value, result_type )
        
        worksheet = self.get_current_sheet()
        worksheet.cell( row = 1, column = 1 ).hyperlink = '#\'' + self.summary_sheet_name + '\'!' + 'A1'
        
        return True
    
    
    def analyze_correl( self ):
       
        if not self.is_target_sheet_set() and self.is_data_loaded():
            return True
            
        worksheet = self.get_current_sheet()
        pearsons_r = self.get_pearsons_r( self.data_list[0], self.data_list[1] )
        worksheet.cell( row = 2, column = 5 ).value = 'Pearson\'s r'
        worksheet.cell( row = 3, column = 5 ).value = pearsons_r            
        worksheet.cell( row = 4, column = 5 ).value = ('High correlation' if pearsons_r > 0.5 else 'Low correlation')
        self.draw_correl_curve( 5, 5 )
            
        return pearsons_r
 
    
    def analyze_anova( self ):
        
        if not self.is_target_sheet_set() and self.is_data_loaded():
            return True
                       
        worksheet = self.get_current_sheet()
                                           
        # set initial position to F3 cell
        rawdata_start_row = 3
        rawdata_start_col = 6        
        longest_data_length = 0        
        
        if self.check_data_type( self.data_list[0] ) == self.type_nom and self.check_data_type( self.data_list[1] ) == self.type_num:
            nom_list = self.data_list[0]
            ratio_list = self.data_list[1]
        elif self.check_data_type( self.data_list[1] ) == self.type_nom and self.check_data_type( self.data_list[0] ) == self.type_num:
            nom_list = self.data_list[1]
            ratio_list = self.data_list[0]            
        else:
            nom_list = []
            ratio_list = []
        
        # suppose list[0] is a list of nominal values
        sole_nominal_list = list( set( nom_list ) )
        sole_nominal_list.sort( reverse = False )        
        
        num_of_nominal_items = len( sole_nominal_list )                
        chart_start_col = rawdata_start_col + num_of_nominal_items + 1        
        
        worksheet.cell( row = rawdata_start_row - 2, column = rawdata_start_col ).value = 'Rawdata'
                
        cur_row = rawdata_start_row
        cur_col = rawdata_start_col

        rawdata_matrix = [] # for anova
        
        # to be changed
        for comp_value in sole_nominal_list:
           
            worksheet.cell( row = rawdata_start_row - 1, column = cur_col ).value = comp_value
            cur_item_index = 0
            item_counter = 0  
            cur_row = rawdata_start_row            
            cur_rawdata_vector = [] # anova
            
            for nominal_value in nom_list:
            
                if nominal_value == comp_value:
                    
                    worksheet.cell( row = cur_row, column = cur_col ).value = ratio_list[cur_item_index]
                    cur_rawdata_vector.append( ratio_list[cur_item_index] ) # for anova
                    cur_row += 1
                    item_counter += 1
                    
                    if longest_data_length < item_counter:
                        longest_data_length = item_counter
                            
                cur_item_index += 1
            
            rawdata_matrix.append( cur_rawdata_vector )
            cur_col += 1
        
        pri_param_start_row = rawdata_start_row + longest_data_length + 3                        
        worksheet.cell( row = pri_param_start_row - 2, column = rawdata_start_col - 1 ).value = 'Primary Parameter Table'
        worksheet.cell( row = pri_param_start_row, column = rawdata_start_col - 1 ).value = 'Min'
        worksheet.cell( row = pri_param_start_row + 1, column = rawdata_start_col - 1 ).value = 'Q1'
        worksheet.cell( row = pri_param_start_row + 2, column = rawdata_start_col - 1 ).value = 'Median'
        worksheet.cell( row = pri_param_start_row + 3, column = rawdata_start_col - 1 ).value = 'Q3'
        worksheet.cell( row = pri_param_start_row + 4, column = rawdata_start_col - 1 ).value = 'Max'
        
        cur_row = pri_param_start_row
        cur_col = rawdata_start_col
        for comp_value in sole_nominal_list:
            
            worksheet.cell( row = pri_param_start_row - 1, column = cur_col ).value = comp_value
            
            rawdata_ref_area = self.get_range_address( cur_col, rawdata_start_row, cur_col, rawdata_start_row + longest_data_length - 1 )
            
            worksheet.cell( row = pri_param_start_row, column = cur_col  ).value = '=_xlfn.MIN(' + rawdata_ref_area + ')'
            worksheet.cell( row = pri_param_start_row + 1, column = cur_col ).value = '=_xlfn.QUARTILE.INC(' + rawdata_ref_area + ',1)'
            worksheet.cell( row = pri_param_start_row + 2, column = cur_col ).value = '=_xlfn.MEDIAN(' + rawdata_ref_area + ')'
            worksheet.cell( row = pri_param_start_row + 3, column = cur_col ).value = '=_xlfn.QUARTILE.INC(' + rawdata_ref_area  + ',3)'
            worksheet.cell( row = pri_param_start_row + 4, column = cur_col ).value = '=_xlfn.MAX(' + rawdata_ref_area + ')'
            
            cur_col += 1
        
        sec_param_start_row = pri_param_start_row + 8
        worksheet.cell( row = sec_param_start_row - 2, column = rawdata_start_col - 1 ).value = 'Secondary Parameter Table'
        worksheet.cell( row = sec_param_start_row, column = rawdata_start_col - 1 ).value = 'Min'
        worksheet.cell( row = sec_param_start_row + 1, column = rawdata_start_col - 1 ).value = 'Q1-Min'
        worksheet.cell( row = sec_param_start_row + 2, column = rawdata_start_col - 1 ).value = 'Med-Q1'
        worksheet.cell( row = sec_param_start_row + 3, column = rawdata_start_col - 1 ).value = 'Q3-Med'
        worksheet.cell( row = sec_param_start_row + 4, column = rawdata_start_col - 1 ).value = 'Max-Q3'
        worksheet.cell( row = sec_param_start_row + 5, column = rawdata_start_col - 1 ).value = 'Mean'            
        
        cur_row = sec_param_start_row
        cur_col = rawdata_start_col
        for comp_value in sole_nominal_list:
            
            worksheet.cell( row = sec_param_start_row - 1, column = cur_col ).value = comp_value      
            rawdata_ref_area = self.get_range_address( cur_col, rawdata_start_row, cur_col, rawdata_start_row + longest_data_length - 1 )                
            worksheet.cell( row = sec_param_start_row, column = cur_col  ).value = '=' + self.get_cell_address( cur_col, pri_param_start_row )
            worksheet.cell( row = sec_param_start_row + 5, column = cur_col ).value = '=_xlfn.AVERAGE(' + rawdata_ref_area + ')'
            
            for cur_index in range(1, 5):
                worksheet.cell( row = sec_param_start_row + cur_index, column = cur_col ).value = '=' +\
                self.get_cell_address( cur_col, pri_param_start_row + cur_index ) + '-' + self.get_cell_address( cur_col, pri_param_start_row + cur_index - 1 )
                
            cur_col += 1
        
        self.draw_boxplot( chart_start_col, rawdata_start_row - 1, rawdata_start_col, sec_param_start_row, num_of_nominal_items )
        
        # to be changed
        num_of_groups = len ( sole_nominal_list )
        num_of_observations = 0
        total_sum = 0        
        for cur_rawdata_vector in rawdata_matrix:            
            cur_vector = numpy.array( cur_rawdata_vector )            
            total_sum += cur_vector.sum()
            num_of_observations += len( cur_vector )
        
        df_of_sst = num_of_observations - 1
        df_of_ssb = num_of_groups - 1
        df_of_ssw = num_of_observations - num_of_groups
        
        if num_of_groups >= num_of_observations or df_of_sst <= 0 or df_of_ssb <= 0:
            return 'BP'
        
        total_mean = total_sum / num_of_observations                        
        ssb = 0
        ssw = 0
        for cur_rawdata_vector in rawdata_matrix:
            cur_vector = numpy.array( cur_rawdata_vector )
            ssb += float( len( cur_vector ) * numpy.power( ( cur_vector.mean() - total_mean ), 2 ) )
            
            for cur_rawdata in cur_rawdata_vector:
                ssw += float( numpy.power( ( cur_rawdata - cur_vector.mean() ), 2 ) )
        
        sst = ssb + ssw        
        msb = ssb / df_of_ssb
        msw = ssw / df_of_ssw
        f_ratio = msb / msw        
        # f_stats = self.wrapper( stats.f_oneway, rawdata_matrix )        
        significance_level = 0.05
        f_critical_value = stats.f.isf( significance_level, dfn = df_of_ssb, dfd = df_of_ssw )        
        p_value = 1 - stats.f.cdf( f_ratio , dfn = df_of_ssb, dfd = df_of_ssw )
                                
        result_start_row = 2
        result_start_col = chart_start_col + 10
        worksheet.cell( row = result_start_row, column = result_start_col ).value = 'Sum of squared total'
        worksheet.cell( row = result_start_row + 1, column = result_start_col  ).value = 'Sum of squared between groups'
        worksheet.cell( row = result_start_row + 2, column = result_start_col  ).value = 'Sum of squared within groups'
        worksheet.cell( row = result_start_row + 3, column = result_start_col  ).value = 'Degree of freedom of SST'
        worksheet.cell( row = result_start_row + 4, column = result_start_col  ).value = 'Degree of freedom of SSB'
        worksheet.cell( row = result_start_row + 5, column = result_start_col  ).value = 'Degree of freedom of SSW'
        worksheet.cell( row = result_start_row + 6, column = result_start_col  ).value = 'Mean squared between groups'
        worksheet.cell( row = result_start_row + 7, column = result_start_col  ).value = 'Mean squared within groups'
        worksheet.cell( row = result_start_row + 8, column = result_start_col  ).value = 'F ratio'
        worksheet.cell( row = result_start_row + 9, column = result_start_col  ).value = 'F critical value'
        worksheet.cell( row = result_start_row + 10, column = result_start_col  ).value = 'Significance level'
        worksheet.cell( row = result_start_row + 11, column = result_start_col  ).value = 'p-value'
        worksheet.column_dimensions[openpyxl.utils.get_column_letter( result_start_col )].width = 30

        worksheet.cell( row = result_start_row, column = result_start_col + 1 ).value = sst
        worksheet.cell( row = result_start_row, column = result_start_col + 1 ).number_format = '0.00'
        worksheet.cell( row = result_start_row + 1, column = result_start_col + 1 ).value = ssb
        worksheet.cell( row = result_start_row + 1, column = result_start_col + 1 ).number_format = '0.00'
        worksheet.cell( row = result_start_row + 2, column = result_start_col + 1 ).value = ssw
        worksheet.cell( row = result_start_row + 2, column = result_start_col + 1 ).number_format = '0.00'
        worksheet.cell( row = result_start_row + 3, column = result_start_col + 1 ).value = df_of_sst
        worksheet.cell( row = result_start_row + 4, column = result_start_col + 1 ).value = df_of_ssb
        worksheet.cell( row = result_start_row + 5, column = result_start_col + 1 ).value = df_of_ssw
        worksheet.cell( row = result_start_row + 6, column = result_start_col + 1 ).value = msb
        worksheet.cell( row = result_start_row + 6, column = result_start_col + 1 ).number_format = '0.00'
        worksheet.cell( row = result_start_row + 7, column = result_start_col + 1 ).value = msw
        worksheet.cell( row = result_start_row + 7, column = result_start_col + 1 ).number_format = '0.00'        
        worksheet.cell( row = result_start_row + 8, column = result_start_col + 1 ).value = f_ratio
        worksheet.cell( row = result_start_row + 8, column = result_start_col + 1 ).number_format = '0.00'        
        worksheet.cell( row = result_start_row + 9, column = result_start_col + 1 ).value = f_critical_value
        worksheet.cell( row = result_start_row + 9, column = result_start_col + 1 ).number_format = '0.00'        
        worksheet.cell( row = result_start_row + 10, column = result_start_col + 1 ).value = significance_level
        worksheet.cell( row = result_start_row + 10, column = result_start_col + 1 ).number_format = '0.00'        
        worksheet.cell( row = result_start_row + 11, column = result_start_col + 1 ).value = p_value        
        worksheet.cell( row = result_start_row + 11, column = result_start_col + 1 ).number_format = '0.00'        
        worksheet.cell( row = result_start_row + 12, column = result_start_col ).value = self.get_hypothesis_test_result( f_ratio, f_critical_value )       
            
        return p_value
      
    
    def analyze_cross_table( self ):
        
        if not self.is_target_sheet_set() and self.is_data_loaded():
            return True
        
        worksheet = self.get_current_sheet()
                                   
        a_list = list( set( self.data_list[0] ) ) # list of items which is listed along the upper row
        a_list.sort( reverse = False )
        b_list = list( set( self.data_list[1] ) ) # list of items which is listed along the left most column
        b_list.sort( reverse = False )                
        
        num_of_a = len( a_list )
        num_of_b = len( b_list )
        
        # set initial positions
        freq_start_row = 3
        freq_end_row = freq_start_row + num_of_b
        prob_start_row = freq_start_row + num_of_b + 4
        exp_freq_start_row = prob_start_row + num_of_b + 4
        chi_square_start_row = exp_freq_start_row + num_of_b + 4        
        every_start_col = 6
        
        # draw tables
        self.draw_cross_table( every_start_col, freq_start_row, freq_end_row, a_list, b_list, 'Frequency' )
        self.draw_cross_table( every_start_col, prob_start_row, freq_end_row, a_list, b_list, 'Probability' )
        self.draw_cross_table( every_start_col, exp_freq_start_row, freq_end_row, a_list, b_list, 'Expected Frequency' )
        self.draw_cross_table( every_start_col, chi_square_start_row, freq_end_row, a_list, b_list, 'Chi Square' )
        
        # analysis result
        result_start_row = 2
        result_start_col = 8 + num_of_a
        worksheet.cell( row = result_start_row, column = result_start_col ).value = 'Rows'
        worksheet.cell( row = result_start_row + 1, column = result_start_col ).value = 'Columns'
        worksheet.cell( row = result_start_row + 2, column = result_start_col ).value = 'Significance level(alpha)'
        worksheet.cell( row = result_start_row + 3, column = result_start_col ).value = 'Degree of freedom'
        worksheet.cell( row = result_start_row + 4, column = result_start_col ).value = 'Critical chi square value'
        worksheet.cell( row = result_start_row + 5, column = result_start_col ).value = 'p-value'
        worksheet.cell( row = result_start_row + 6, column = result_start_col ).value = 'Total number of observation'
        worksheet.cell( row = result_start_row + 7, column = result_start_col ).value = 'Pearson chi square'
        worksheet.cell( row = result_start_row + 8, column = result_start_col ).value = 'Phi coefficient'
        worksheet.cell( row = result_start_row + 9, column = result_start_col ).value = 'Cramer\'s V'
        worksheet.column_dimensions[openpyxl.utils.get_column_letter( result_start_col )].width = 30
        
        pearson_chi_square = self.get_pearson_chi_square( self.data_list[0], self.data_list[1] )
        significance_level = 0.05
        degree_of_freedom = ( num_of_a - 1 ) * ( num_of_b - 1 )
        critical_chi_square = stats.chi2.isf( q = significance_level, df = degree_of_freedom  ) # to be changed        
        total_number = min( len( self.data_list[0] ), len( self.data_list[1] ) )
        p_value = 1 - stats.chi2.cdf( pearson_chi_square, degree_of_freedom ) # to be changed
        phi_coefficient = pearson_chi_square / total_number
        cramers_v = math.sqrt( phi_coefficient / min( num_of_a, num_of_b )  ) 
        
        worksheet.cell( row = result_start_row, column = result_start_col + 1 ).value = num_of_a
        worksheet.cell( row = result_start_row + 1, column = result_start_col + 1 ).value = num_of_b
        worksheet.cell( row = result_start_row + 2, column = result_start_col + 1 ).value = significance_level        
        worksheet.cell( row = result_start_row + 2, column = result_start_col + 1 ).number_format = '0.00'
        worksheet.cell( row = result_start_row + 3, column = result_start_col + 1 ).value = degree_of_freedom        
        worksheet.cell( row = result_start_row + 4, column = result_start_col + 1 ).value = critical_chi_square        
        worksheet.cell( row = result_start_row + 4, column = result_start_col + 1 ).number_format = '0.00'
        worksheet.cell( row = result_start_row + 5, column = result_start_col + 1 ).value = p_value        
        worksheet.cell( row = result_start_row + 5, column = result_start_col + 1 ).number_format = '0.00'
        worksheet.cell( row = result_start_row + 6, column = result_start_col + 1 ).value = total_number        
        worksheet.cell( row = result_start_row + 7, column = result_start_col + 1 ).value = pearson_chi_square        
        worksheet.cell( row = result_start_row + 7, column = result_start_col + 1 ).number_format = '0.00'
        worksheet.cell( row = result_start_row + 8, column = result_start_col + 1 ).value = phi_coefficient        
        worksheet.cell( row = result_start_row + 8, column = result_start_col + 1 ).number_format = '0.00'
        worksheet.cell( row = result_start_row + 9, column = result_start_col + 1 ).value = cramers_v
        worksheet.cell( row = result_start_row + 9, column = result_start_col + 1 ).number_format = '0.00'                    
        worksheet.cell( row = result_start_row + 10, column = result_start_col ).value = self.get_hypothesis_test_result( pearson_chi_square, critical_chi_square )       
        
        return p_value
  
         
    # ===== functions to draw curves or tables =====
    
    def draw_correl_curve( self, chart_col, chart_row ):
        
        if not self.is_target_sheet_set() and self.is_data_loaded():
            return True
          
        worksheet = self.get_current_sheet()
        correl_chart = openpyxl.chart.ScatterChart()            
        correl_chart.style = 18 # to be changed            
        correl_chart.legend = None                        
        x_ref = openpyxl.chart.Reference(worksheet, min_col = 2, max_col = 2, min_row = self.start_row, max_row = self.end_row ) # to be changed
        y_ref = openpyxl.chart.Reference(worksheet, min_col = 3, max_col = 3, min_row = self.start_row, max_row = self.end_row ) # to be changed
    
        correl_series = openpyxl.chart.Series(y_ref, x_ref, title_from_data = False )
        correl_series.marker = openpyxl.chart.marker.Marker( symbol = 'circle' )            
        correl_series.graphicalProperties.line.noFill = True
        correl_trendline = openpyxl.chart.trendline.Trendline( trendlineType = 'linear' )                                               
        correl_trendline.spPr = openpyxl.chart.shapes.GraphicalProperties( ln = openpyxl.drawing.line.LineProperties( w = 20000, solidFill = self.color_light_blue ) )
        correl_series.trendline = correl_trendline
        correl_chart.series.append( correl_series ) 
    
        worksheet.add_chart( correl_chart, self.get_cell_address( chart_col, chart_row ) )
        
        return True
  
    
    def draw_boxplot( self, chart_col, chart_row, data_range_origin_col, data_range_origin_row, num_of_items ):
        
        if not self.is_target_sheet_set() and self.is_data_loaded():
            return True
        
        data_range_end_col = data_range_origin_col + num_of_items - 1
        
        worksheet = self.get_current_sheet()
        
        # generate charts
        boxplot_chart = openpyxl.chart.BarChart()                    
        boxplot_chart.legend = None
        boxplot_chart.grouping = 'stacked'
        boxplot_chart.type = 'col'
        boxplot_chart.overlap = 100 # to set on the same axis
        boxplot_chart.height = 11 # to be changed
        
        boxplot_mean_chart = openpyxl.chart.LineChart()
        # boxplot_mean_chart.style = 18 # to be changed
        boxplot_mean_chart.legend = None
        
        # references
        item_ref = openpyxl.chart.Reference(worksheet, min_col = data_range_origin_col, max_col = data_range_end_col,\
                                            min_row = data_range_origin_row - 1, max_row = data_range_origin_row - 1 )
        min_ref = openpyxl.chart.Reference(worksheet, min_col = data_range_origin_col, max_col = data_range_end_col,\
                                           min_row = data_range_origin_row, max_row = data_range_origin_row )
        q1_minus_min_ref = openpyxl.chart.Reference(worksheet, min_col = data_range_origin_col, max_col = data_range_end_col,\
                                                    min_row = data_range_origin_row + 1, max_row = data_range_origin_row + 1 )
        med_minus_q1_ref = openpyxl.chart.Reference(worksheet, min_col = data_range_origin_col, max_col = data_range_end_col,\
                                                    min_row = data_range_origin_row + 2, max_row = data_range_origin_row + 2 )
        q3_minus_med_ref = openpyxl.chart.Reference(worksheet, min_col = data_range_origin_col, max_col = data_range_end_col,\
                                                    min_row = data_range_origin_row + 3, max_row = data_range_origin_row + 3 )
        max_minus_q3_ref = openpyxl.chart.Reference(worksheet, min_col = data_range_origin_col, max_col = data_range_end_col,\
                                                    min_row = data_range_origin_row + 4, max_row = data_range_origin_row + 4)
        mean_ref = openpyxl.chart.Reference(worksheet, min_col = data_range_origin_col, max_col = data_range_end_col,\
                                            min_row = data_range_origin_row + 5, max_row = data_range_origin_row + 5)
        
        # Series
        min_series = openpyxl.chart.Series( min_ref, title_from_data = False )
        q1_minus_min_series = openpyxl.chart.Series( q1_minus_min_ref, title_from_data = False )
        med_minus_q1_series = openpyxl.chart.Series( med_minus_q1_ref, title_from_data = False )                
        q3_minus_med_series = openpyxl.chart.Series( q3_minus_med_ref, title_from_data = False )
        max_minus_q3_series = openpyxl.chart.Series( max_minus_q3_ref, title_from_data = False )
        
        max_err_datasource = openpyxl.chart.data_source.NumDataSource( numRef = openpyxl.chart.data_source.NumRef( max_minus_q3_ref ) )
        q1_minus_min_series.errBars = openpyxl.chart.error_bar.ErrorBars( errBarType = 'minus', errValType = 'percentage', val = 100 ) 
        q3_minus_med_series.errBars = openpyxl.chart.error_bar.ErrorBars( errBarType = 'plus', errValType = 'cust', plus = max_err_datasource )
        
        no_fill_bar_gp = openpyxl.chart.shapes.GraphicalProperties( noFill = True )
        
        min_series.spPr = no_fill_bar_gp        
        q1_minus_min_series.spPr = no_fill_bar_gp
        med_minus_q1_series.graphicalProperties.solidFill = self.color_orange_red # red
        q3_minus_med_series.graphicalProperties.solidFill = self.color_heavy_blue        
        max_minus_q3_series.spPr = no_fill_bar_gp
        
        mean_series = openpyxl.chart.Series( mean_ref, title_from_data = False )
        mean_series.graphicalProperties.line.noFill = True
        mean_series.marker = openpyxl.chart.marker.Marker( symbol = 'x' )
        mean_series.marker.graphicalProperties.line.solidFill = '000000'
                
        min_series.cat = openpyxl.chart.data_source.AxDataSource( numRef = openpyxl.chart.data_source.NumRef( item_ref ) ) 
        q1_minus_min_series.cat = openpyxl.chart.data_source.AxDataSource( numRef = openpyxl.chart.data_source.NumRef( item_ref ) ) 
        med_minus_q1_series.cat = openpyxl.chart.data_source.AxDataSource( numRef = openpyxl.chart.data_source.NumRef( item_ref ) ) 
        q3_minus_med_series.cat = openpyxl.chart.data_source.AxDataSource( numRef = openpyxl.chart.data_source.NumRef( item_ref ) ) 
        max_minus_q3_series.cat = openpyxl.chart.data_source.AxDataSource( numRef = openpyxl.chart.data_source.NumRef( item_ref ) ) 
        mean_series.cat = openpyxl.chart.data_source.AxDataSource( numRef = openpyxl.chart.data_source.NumRef( item_ref ) ) 
        
        # add series to chart
        boxplot_chart.series.append( min_series ) 
        boxplot_chart.series.append( q1_minus_min_series ) 
        boxplot_chart.series.append( med_minus_q1_series ) 
        boxplot_chart.series.append( q3_minus_med_series ) 
        boxplot_chart.series.append( max_minus_q3_series ) 
        boxplot_mean_chart.series.append( mean_series )
        boxplot_chart += boxplot_mean_chart        
        
        # add chart to sheet
        worksheet.add_chart( boxplot_chart, self.get_cell_address( chart_col, chart_row ) )
        
        return True
    
    
    def draw_cross_table( self, start_col, start_row, freq_end_row, a_list, b_list, title ):
        
        if not self.is_target_sheet_set() and self.is_data_loaded():
            return True
        
        worksheet = self.get_current_sheet()
        
        num_of_a = len( a_list )
        num_of_b = len( b_list )        
        
        end_row = start_row + num_of_b
        end_col = start_col + num_of_a       
        
        worksheet.cell( row = start_row - 2, column = start_col - 1 ).value = title
        worksheet.cell( row = start_row - 1, column = start_col - 1 ).value = 'Item'
        worksheet.cell( row = end_row, column = end_col ).value = self.get_sum_formula( self.get_range_address( end_col, start_row, end_col, end_row - 1 ) )                    
        worksheet.cell( row = end_row, column = end_col ).number_format = '0.00'
        
        # upper row setting
        for cur_index in range( 0, num_of_a ):
            worksheet.cell( row = start_row - 1, column = start_col + cur_index ).value = a_list[cur_index]
                    
        worksheet.cell( row = start_row - 1, column = end_col ).value = 'Row Sum'
        
        # lower row setting
        for cur_index in range( 0, num_of_a ):
            cur_col = start_col + cur_index            
            worksheet.cell( row = end_row, column = cur_col ).value = self.get_sum_formula( self.get_range_address( cur_col, start_row, cur_col, end_row - 1 ) )            
            worksheet.cell( row = end_row, column = cur_col ).number_format = '0.00'
                
        # left most column setting
        for cur_index in range( 0, num_of_b ):
            worksheet.cell( row = start_row + cur_index, column = start_col - 1 ).value = b_list[cur_index]
            
        worksheet.cell( row = end_row, column = start_col - 1 ).value = 'Col Sum'        
        
        # right most column setting
        for cur_index in range( 0, num_of_b ):
            cur_row = start_row + cur_index            
            worksheet.cell( row = cur_row, column = end_col ).value = self.get_sum_formula( self.get_range_address( start_col, cur_row, end_col - 1, cur_row ) )            
            worksheet.cell( row = cur_row, column = end_col ).number_format = '0.00'
        
        # filling inside table
        for cur_a_index in range(0, num_of_a ):                       
            
            cur_col = start_col + cur_a_index
            
            for cur_b_index in range(0, num_of_b ):
                
                cur_row = start_row + cur_b_index
                formula = ''                
                
                if title == 'Frequency':
                    a_ref_range = self.get_range_address( self.start_col, self.start_row, self.start_col, self.end_row )
                    a_criteria_cell = self.get_cell_address( cur_col, start_row - 1)
                    b_ref_range = self.get_range_address( self.end_col, self.start_row, self.end_col, self.end_row )
                    b_criteria_cell = self.get_cell_address( start_col - 1, cur_row )
                    formula = self.get_freq_formula( a_ref_range, a_criteria_cell, b_ref_range, b_criteria_cell )
                    
                elif title == 'Probability':
                    dividend_cell = self.get_cell_address( cur_col, cur_row - num_of_b - 4 )
                    divisor_cell = self.get_cell_address( end_col, freq_end_row )
                    formula = self.get_prob_formula( dividend_cell, divisor_cell )
                    
                elif title == 'Expected Frequency':
                    total_cell = self.get_cell_address( end_col, freq_end_row )
                    a_prob_cell = self.get_cell_address( cur_col, end_row - num_of_b - 4 )
                    b_prob_cell = self.get_cell_address(  end_col, cur_row - num_of_b - 4)
                    formula = self.get_exp_freq_formula( total_cell, a_prob_cell, b_prob_cell )
                
                elif title == 'Chi Square':                    
                    freq_cell = self.get_cell_address( cur_col, cur_row - 3 * ( num_of_b ) - 12 )
                    exp_freq_cell = self.get_cell_address(cur_col, cur_row - num_of_b - 4 )
                    formula = self.get_chi_square_formula( freq_cell, exp_freq_cell )
                
                else:
                    formula == ''
                    
                worksheet.cell( row = cur_row, column = cur_col ).value = formula
                worksheet.cell( row = cur_row, column = cur_col ).number_format = '0.00'
                
        return True
    